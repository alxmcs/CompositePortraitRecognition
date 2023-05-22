import os
import sqlite3

import numpy as np
import openpyxl
from sklearn.decomposition import PCA
from sklearn.discriminant_analysis import QuadraticDiscriminantAnalysis
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler

import common.db.db_operations as db

CLF = QuadraticDiscriminantAnalysis()

QUERIES = {
    'same': """select value_p, value_s, pt_id, ? as class  
        from 
        (select value as value_p, person_id as pt_id from embedding where model_id = ? and info = ?)
        inner join
        (select value as value_s, person_id as st_id from embedding where model_id = ? and info = ?)
        on pt_id = st_id 
        where value_p is not null and value_s is not null 
        order by pt_id asc""",
    'different': """select value_p, value_s, pt_id, ? as class 
        from 
        (select value as value_p, person_id as pt_id from embedding where model_id = ? and info = ?)
        inner join
        (select value as value_s, person_id as st_id from embedding where model_id = ? and info = ?)
        on pt_id = st_id 
        where value_p is not null and value_s is not null 
        order by pt_id asc"""
}

models = [
    "VGG-Face",
    "Facenet",
    "Facenet512",
    "OpenFace",
    "DeepFace",
    "DeepID",
    "ArcFace",
]


def get_preprocessed_data(db_data):
    for row in db_data:
        row[0] = row[0][1:-1]
        row[0] = np.fromstring(row[0], dtype=float, sep=', ')
        row[1] = row[1][1:-1]
        row[1] = np.fromstring(row[1], dtype=float, sep=', ')
        row[0] -= row[1]
    return [np.array([db_data[x][0] for x in range(0, len(db_data))]),
            np.array([db_data[x][2] for x in range(0, len(db_data))], dtype=int),
            np.array([db_data[x][3] for x in range(0, len(db_data))], dtype=int)]


def get_data(cursor, key, params):
    db_data = [list(item) for item in cursor.execute(QUERIES[key], params).fetchall()]
    return get_preprocessed_data(db_data)


def test_data(same_data, diff_data, accuracy_array, precision_array, recall_array, f1_array):
    data = np.concatenate((same_data[0], diff_data[0]))
    target = np.concatenate((same_data[2], diff_data[2]))
    data = StandardScaler().fit_transform(data)
    # print(data.shape)
    # print(target.shape)

    pca_ = 0.4
    pca = PCA(pca_)
    pca.fit(data)
    data = pca.transform(data)
    x_train, x_test, y_train, y_test = train_test_split(data, target, test_size=0.3, random_state=0,
                                                        stratify=target)
    CLF.fit(x_train, y_train)
    y_pred = CLF.predict(x_test)
    accuracy_array.append(accuracy_score(y_test, y_pred))
    precision_array.append(precision_score(y_test, y_pred))
    recall_array.append(recall_score(y_test, y_pred))
    f1_array.append(f1_score(y_test, y_pred))


def train_classifier():
    photo_true_st = f'photo_true_Facenet512_tdcs_st'
    photo_false_st = f'photo_false_Facenet512_tdcs_st'
    sketch_true = f'sketch_true_Facenet512_tdcs'
    sketch_false = f'sketch_false_Facenet512_tdcs'

    conn = sqlite3.connect("../../common/db/database.db")
    cursor = conn.cursor()

    same = 1
    diff = 0
    model_id = 5

    same_data_st = get_data(cursor, 'same', [same, model_id, photo_true_st, model_id, sketch_true])
    diff_data_st = get_data(cursor, 'different', [diff, model_id, photo_false_st, model_id, sketch_false])

    data = np.concatenate((same_data_st[0], diff_data_st[0]))
    target = np.concatenate((same_data_st[2], diff_data_st[2]))

    data = StandardScaler().fit_transform(data)
    pca_ = 0.4
    pca = PCA(pca_)
    pca.fit(data)
    data = pca.transform(data)
    CLF.fit(data, target)
    return CLF, pca


if __name__ == "__main__":
    db_path = os.path.join('C:\\CompositePortraitRecongnition', 'common', 'db', 'database.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    same = 1
    diff = 0
    ST = False
    dataset_name = 'tdcs'

    accuracy_array = []
    precision_array = []
    recall_array = []
    f1_array = []

    accuracy_array_st = []
    precision_array_st = []
    recall_array_st = []
    f1_array_st = []

    accuracy_array_gray = []
    precision_array_gray = []
    recall_array_gray = []
    f1_array_gray = []

    for i in range(0, len(models)):
        model_name = models[i]
        photo_true = f'photo_true_{model_name}_{dataset_name}'
        sketch_true = f'sketch_true_{model_name}_{dataset_name}'
        sketch_false = f'sketch_false_{model_name}_{dataset_name}'
        photo_true_st = f'photo_true_{model_name}_{dataset_name}_st1'
        photo_false_st = f'photo_false_{model_name}_{dataset_name}_st1'
        gray_photo_true = f'gray_photo_true_{model_name}_{dataset_name}'
        gray_sketch_true = f'gray_sketch_true_{model_name}_{dataset_name}'
        gray_sketch_false = f'gray_sketch_false_{model_name}_{dataset_name}'

        model_id = cursor.execute(db.QUERIES['get_model_id_by_name'], [model_name]).fetchone()[0]

        same_data_st = get_data(cursor, 'same', [same, model_id, photo_true_st, model_id, sketch_true])
        diff_data_st = get_data(cursor, 'different', [diff, model_id, photo_false_st, model_id, sketch_false])

        same_data = get_data(cursor, 'same', [same, model_id, photo_true, model_id, sketch_true])
        diff_data = get_data(cursor, 'different', [diff, model_id, photo_true, model_id, sketch_false])

        gray_same_data = get_data(cursor, 'same', [same, model_id, gray_photo_true, model_id, gray_sketch_true])
        gray_diff_data = get_data(cursor, 'different', [diff, model_id, gray_photo_true, model_id, gray_sketch_false])

        test_data(same_data, diff_data, accuracy_array, precision_array, recall_array, f1_array)
        test_data(same_data_st, diff_data_st, accuracy_array_st, precision_array_st, recall_array_st, f1_array_st)
        test_data(gray_same_data, gray_diff_data, accuracy_array_gray, precision_array_gray, recall_array_gray,
                  f1_array_gray)

    book = openpyxl.Workbook()
    sheet_1 = book.create_sheet("results", 0)
    sheet_2 = book.create_sheet("results with ST", 1)
    sheet_3 = book.create_sheet("results gray", 2)
    headers = ['backbone', 'accuracy', 'precision', 'recall', 'f1']
    sheet_1.append(headers)
    sheet_2.append(headers)
    sheet_3.append(headers)
    for name, ac, pr, rec, f1 in zip(models, accuracy_array, precision_array, recall_array,
                                     f1_array):
        print(f'model name = {name} accuracy = {ac:.2f}, precision = {pr:.2f}, '
              f'recall = {rec:.2f}', f'score = {f1:.2f}')
        sheet_1.append([name, ac, pr, rec, f1])
    for name, ac, pr, rec, f1 in zip(models, accuracy_array_gray, precision_array_gray, recall_array_gray,
                                     f1_array_gray):
        sheet_3.append([name, ac, pr, rec, f1])
    for name, ac, pr, rec, f1 in zip(models, accuracy_array_st, precision_array_st, recall_array_st,
                                     f1_array_st):
        sheet_2.append([name, ac, pr, rec, f1])
    book.save("results.xlsx")
